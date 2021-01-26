
''' module with utilities for excel
'''
import io
import re
from openpyxl.styles import Border, Side, PatternFill
from openpyxl.utils.cell import get_column_letter, column_index_from_string
from openpyxl.formatting.rule import CellIsRule


def set_vertical_cells(ws, loc_init: str, values: list, font, alignment):
    # loc_init: format "A0" multiple capital letters followd by multiple numbers
    loc = re.match(r'^([A-Z]+)([0-9]+)$', loc_init)
    row_init = int(loc.group(2))
    col_init = loc.group(1)
    for i, value in enumerate(values):
        loc = col_init + str(row_init + i)
        cell = ws[loc]
        cell.value = value
        cell.font = font
        cell.alignment = alignment


def set_horizontal_cells(ws, loc_init: str, values: list, font, alignment):
    # loc_init: format "A0" multiple capital letters followd by multiple numbers
    loc = re.match(r'^([A-Z]+)([0-9]+)$', loc_init)
    row_init = int(loc.group(2))
    col_init_num = column_index_from_string(loc.group(1))
    for i, value in enumerate(values):
        loc = get_column_letter(col_init_num + i) + str(row_init)
        cell = ws[loc]
        cell.value = value
        cell.font = font
        cell.alignment = alignment


def format_vertical(ws, loc_range: str, num_format: str):
    # loc_range: format "A10:D10" multiple capital letters followd by multiple numbers
    loc = re.match(r'^([A-Z]+)([0-9]+):([A-Z]+)([0-9]+)$', loc_range)
    row_start_num = int(loc.group(2))
    row_end_num = int(loc.group(4))
    col_init = loc.group(1)
    for row in range(row_start_num, row_end_num + 1):
        loc = col_init + str(row)
        ws[loc].number_format = num_format


def format_horizontal(ws, loc_range: str, num_format: str):
    # loc_range: format "A10:D10" multiple capital letters followd by multiple numbers
    loc = re.match(r'^([A-Z]+)([0-9]+):([A-Z]+)([0-9]+)$', loc_range)
    row_init = int(loc.group(2))
    col_start_num = column_index_from_string(loc.group(1))
    col_end_num = column_index_from_string(loc.group(3))
    for col in range(col_start_num, col_end_num + 1):
        loc = get_column_letter(col) + str(row_init)
        ws[loc].number_format = num_format


def conditional_format(ws, cell_loc: str, format_vals: tuple, colors: tuple):
    value = ws[cell_loc].value
    if not isinstance(value, (int, float)):
        return

    if value < format_vals[0]:
        ws.conditional_formatting.add(
            cell_loc, CellIsRule(operator='lessThan',
            formula=[format_vals[0]], fill=PatternFill(bgColor=colors[0])))

    elif format_vals[0] <= value < format_vals[1]:
        ws.conditional_formatting.add(
            cell_loc, CellIsRule(operator='lessThan',
            formula=[format_vals[1]],
            fill=PatternFill(bgColor=colors[1])))

    else:
        ws.conditional_formatting.add(
            cell_loc, CellIsRule(operator='greaterThan',
            formula=[format_vals[1]], fill=PatternFill(bgColor=colors[2])))


def set_outer_border(ws, cell_range, style='thin'):
    cells =ws[cell_range]
    row = None
    for i, row in enumerate(cells):
        if i == 0:
            for col in row:
                col.border = col.border + Border(top=Side(style=style))

        row[0].border = row[0].border + Border(left=Side(style=style))
        row[-1].border = row[-1].border + Border(right=Side(style=style))

    for col in row:
        col.border = col.border + Border(bottom=Side(style=style))


def set_border(ws, cell_range, style='thin'):
    border = Border(
        left=Side(border_style=style), right=Side(border_style='thin'),
        top=Side(border_style='thin'), bottom=Side(border_style='thin'),
    )
    rows = ws[cell_range]
    for row in  rows:
        for cell in row:
            cell.border = border


def set_color(ws, cell_range, color=None):
    color_fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.fill = color_fill


def set_column_widths(ws, start_column: str, widths: list):
    col_num = column_index_from_string(start_column)
    for width in widths:
        loc = get_column_letter(col_num)
        ws.column_dimensions[loc].width = width
        col_num += 1


def save_excel(wb):
    ''' method to save excel data to a BytesIO buffer
    '''
    f_excel = io.BytesIO()
    wb.save(f_excel)
    f_excel.seek(0)
    return f_excel


def get_row_column(colrow: str):
    loc = re.match(r'^([A-Z]+)([0-9]+)$', colrow)
    row = int(loc.group(2))
    col = loc.group(1)
    return row, col
