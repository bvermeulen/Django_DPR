''' module to generate excel version of mpr
'''
import calendar
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from daily_report.report_backend import ReportInterface
from seismicreport.utils.utils_funcs import calc_ratio
from seismicreport.utils.utils_excel import (
    set_vertical_cells, set_horizontal_cells, format_vertical, format_horizontal,
    set_column_widths, set_color, save_excel,
)
from seismicreport.vars import (
    CONTRACT, TCF_table, source_prod_schema, time_breakdown_schema, hse_weather_schema
)

fontname = 'Tahoma'
font_large_bold = Font(name=fontname, bold=True, size=11)
font_normal = Font(name=fontname, size=9)
font_bold = Font(name=fontname, bold=True, size=9)
lightblue = 'D8EFF8'

#pylint: disable=line-too-long


class ExcelMprReport():
    ''' class to create excel mpr report
    '''
    def __init__(self, day):
        self.day = day
        self.sourcetype_names = [
            stype.sourcetype_name for stype in self.day.project.sourcetypes.all()
        ]
        self.wb = Workbook()
        self.wb.remove_sheet(self.wb.get_sheet_by_name('Sheet'))
        self.ws = self.wb.create_sheet('MPR')
        self.ws.sheet_view.showGridLines = False

        self.ws_stypes = {}
        if len(self.sourcetype_names) > 1:
            for stype_name in self.sourcetype_names:
                self.ws_stypes[stype_name] = self.wb.create_sheet(stype_name)
                self.ws_stypes[stype_name].sheet_view.showGridLines = False

        self.wp = self.wb.create_sheet('Project')
        self.wp.sheet_view.showGridLines = False

        r_iface = ReportInterface('')
        _, self.prod_total, self.times_total, _, _ = r_iface.calc_totals(self.day)
        (
            self.prod_series_by_type, self.prod_series,
            self.time_series, _, self.hse_series
        ) = r_iface.series

    def get_parameters(self, prod_total, times_total, header_sourcetype):
        params = {}
        params['project'] = self.day.project.project_name
        params['crew'] = self.day.project.crew_name
        params['month'] = self.day.production_date.month
        params['year'] = self.day.production_date.year
        params['days'] = self.day.production_date.day
        params['sourcetype_name'] = header_sourcetype['name']
        params['sweep'] = header_sourcetype['sweep']
        params['vibes'] = header_sourcetype['vibes']
        params['moveup'] = header_sourcetype['moveup']
        params['rechours'] = header_sourcetype['rechours']
        params['rate'] = header_sourcetype['rate']
        params['production_days'] = (
            self.day.production_date - self.day.project.planned_start_date).days + 1
        return params

    def collate_mpr_data(self, prod_series, time_series, hse_series):
        prod_keys = ['date_series']
        prod_keys += [f'{key[:5]}_series' for key in source_prod_schema]
        prod_keys += ['total_sp_series', 'tcf_series', 'ctm_series', 'appctm_series']
        p_series = {f'{key[:-7]}':prod_series[key] for key in prod_keys}

        time_keys = [f'{key}_series' for key in time_breakdown_schema]
        time_keys += [
            'ops_series', 'standby_series', 'downtime_series', 'total_time_series']
        t_series = {f'{key[:-7]}': time_series[key] for key in time_keys}
        proj_df = pd.DataFrame({**p_series, **t_series})

        if hse_series:
            hse_keys = ['date_series']
            hse_keys += [f'{key}_series' for key in hse_weather_schema[:12]]
            h_series = {f'{key[:-7]}': hse_series[key] for key in hse_keys}
            hse_df = pd.DataFrame(h_series)
            proj_df = proj_df.merge(hse_df, how='left', left_on='date', right_on='date')

        proj_df['date'] = pd.to_datetime(proj_df['date'])

        return proj_df

    def create_tab_mpr(self, ws, params, proj_df):
        ''' method to create excel mpr report tab
        '''
        def calc_mpr_values():
            # calculate month values
            m_df = proj_df[
                (proj_df.date.dt.month == params['month']) &
                (proj_df.date.dt.year == params['year'])
            ]

            dates = m_df.date.dt.strftime('%d-%b-%y').to_list()
            m_day = params['days'] - 1
            prod_days = [
                params['production_days'] - m_day + d for d in range(m_day + 1)]

            vp_type = {}
            vp_typesum = {}
            vp_type['total'] = np.array(m_df.total_sp.to_list(), dtype=float)
            vp_type['cumtotal'] = np.cumsum(vp_type['total'])
            vp_type['tcf'] = np.array(m_df.tcf.to_list(), dtype=float)
            vp_type['ctm'] = np.array(m_df.ctm.to_list(), dtype=float)
            vp_type['appctm'] = np.array(m_df.appctm.to_list(), dtype=float)
            vp_typesum['total'] = np.sum(vp_type['total'])
            vp_typesum['ctm'] = np.nansum(vp_type['ctm'])
            vp_typesum['appctm'] = calc_ratio(vp_typesum['total'], vp_typesum['ctm'])

            # calculate the vp terrain distributions and sums
            vp_dist = {}
            vp_distsum = {}
            vp_typesum['tcf'] = 0
            for col in ['sp_t1', 'sp_t2', 'sp_t3', 'sp_t4', 'sp_t5', 'skips']:
                vp_type[col] = np.array(m_df[col].to_list(), dtype=float)
                vp_typesum[col] = np.sum(vp_type[col])
                vp_dist[col] = np.divide(
                    vp_type[col], vp_type['total'], out=np.zeros_like(
                        vp_type[col]), where=vp_type['total']!=0)

                if vp_typesum['total'] > 0:
                    vp_distsum[col] = vp_typesum[col] / vp_typesum['total']
                    if col != 'skips':
                        vp_typesum['tcf'] += vp_distsum[col] * TCF_table[col]

                else:
                    vp_distsum[col] = 0

            # calculate times and sums
            tm_type = {}
            tm_typesum = {}
            for col in ['rec_hours', 'ops', 'standby', 'downtime']:
                tm_type[col] = np.array(m_df[col].to_list())
                tm_typesum[col] = np.sum(tm_type[col])

            tm_type['other_ops'] = tm_type['ops'] - tm_type['rec_hours']
            tm_typesum['other_ops'] = np.sum(tm_type['other_ops'])

            if params['rechours']:
                tm_type['padtime'] = tm_type['rec_hours'] / params['rechours']
                tm_typesum['padtime'] = (
                    tm_typesum['rec_hours'] / (params['rechours'] * params['days']))

            else:
                tm_type['padtime'] = tm_type['rec_hours'] * 0
                tm_typesum['padtime'] = 0

            return (
                dates, prod_days, vp_type, vp_typesum, vp_dist, vp_distsum,
                tm_type, tm_typesum
            )

        (
            dates, prod_days, vp_type, vp_typesum, vp_dist, vp_distsum,
            tm_type, tm_typesum
        ) = calc_mpr_values()

        # set the titles
        ws.merge_cells('I1:M1')
        ws['I1'].value = 'Production Record and Bonus Calculator'
        ws['I1'].font = font_large_bold

        for r in range(2, 12):
            ws.merge_cells(f'A{r}:C{r}')
            ws.merge_cells(f'D{r}:E{r}')

        monthyear = f'{calendar.month_name[params["month"]]}, {params["year"]}'

        set_vertical_cells(ws, 'A2', ['Contract'], font_bold, Alignment(horizontal='left'))
        set_vertical_cells(ws, 'A3', ['Project'], font_bold, Alignment(horizontal='left'))
        set_vertical_cells(ws, 'A4', ['Crew'], font_bold, Alignment(horizontal='left'))
        set_vertical_cells(ws, 'A5', ['Month, Year'], font_bold, Alignment(horizontal='left'))
        set_vertical_cells(ws, 'A6', ['Days in month'], font_bold, Alignment(horizontal='left'))
        set_vertical_cells(ws, 'A7', ['Sourcetype name'], font_bold, Alignment(horizontal='left'))
        set_vertical_cells(ws, 'A8', ['Sweep length'], font_bold, Alignment(horizontal='left'))
        set_vertical_cells(ws, 'A9', ['Contractual vibrators'], font_bold, Alignment(horizontal='left'))
        set_vertical_cells(ws, 'A10', ['Contractual recording hours'], font_bold, Alignment(horizontal='left'))
        set_vertical_cells(ws, 'A11', ['Move up time'], font_bold, Alignment(horizontal='left'))

        set_vertical_cells(ws, 'D2', [CONTRACT], font_normal, Alignment(horizontal='right'))
        set_vertical_cells(ws, 'D3', [f'{params["project"]}'], font_normal, Alignment(horizontal='right'))
        set_vertical_cells(ws, 'D4', [f'{params["crew"]}'], font_normal, Alignment(horizontal='right'))
        set_vertical_cells(ws, 'D4', [f'{params["crew"]}'], font_normal, Alignment(horizontal='right'))
        set_vertical_cells(ws, 'D5', [monthyear], font_normal, Alignment(horizontal='right'))
        set_vertical_cells(ws, 'D6', [params['days']], font_normal, Alignment(horizontal='right'))
        set_vertical_cells(ws, 'D7', [params['sourcetype_name']], font_normal, Alignment(horizontal='right'))
        set_vertical_cells(ws, 'D8', [params['sweep']], font_normal, Alignment(horizontal='right'))
        set_vertical_cells(ws, 'D9', [params['vibes']], font_normal, Alignment(horizontal='right'))
        set_vertical_cells(ws, 'D10', [params['rechours']], font_normal, Alignment(horizontal='right'))
        set_vertical_cells(ws, 'D11', [params['moveup']], font_normal, Alignment(horizontal='right'))

        for r in range(5, 11):
            ws.merge_cells(f'H{r}:J{r}')
            ws.merge_cells(f'K{r}:L{r}')

        set_vertical_cells(ws, 'H5', ['Terrain correction factor'], font_bold, Alignment(horizontal='left'))
        set_vertical_cells(ws, 'H6', ['CTM month'], font_bold, Alignment(horizontal='left'))
        set_vertical_cells(ws, 'H7', ['APP month'], font_bold, Alignment(horizontal='left'))
        set_vertical_cells(ws, 'H8', ['APP / CTM'], font_bold, Alignment(horizontal='left'))
        set_vertical_cells(ws, 'H9', ['Standby days for month'], font_bold, Alignment(horizontal='left'))
        set_vertical_cells(ws, 'H10', ['Effective month rate'], font_bold, Alignment(horizontal='left'))

        set_vertical_cells(ws, 'K5', [vp_typesum['tcf']], font_normal, Alignment(horizontal='right'))
        set_vertical_cells(ws, 'K6', [vp_typesum['ctm']], font_normal, Alignment(horizontal='right'))
        set_vertical_cells(ws, 'K7', [vp_typesum['total']], font_normal, Alignment(horizontal='right'))
        set_vertical_cells(ws, 'K8', [vp_typesum['appctm']], font_normal, Alignment(horizontal='right'))
        set_vertical_cells(ws, 'K9', [tm_typesum['standby'] / 24], font_normal, Alignment(horizontal='right'))
        set_vertical_cells(ws, 'K10', [params['rate']], font_normal, Alignment(horizontal='right'))

        # and some formatting ...
        for r in [5, 10]:
            ws[f'K{r}'].number_format = '0.00%'

        for r in [8, 9]:
            ws[f'K{r}'].number_format = '0.000'

        for r in [6, 7]:
            ws[f'K{r}'].number_format = '#,##0'

        # set the table
        header = [
            'Date', 'Prod. day', 'Flat', 'Rough', 'Facilities', 'Dunes', 'Sabkha',
            'Skips', 'Total VP', 'Cum VPs', 'Rec. hours', 'Other Ops', 'Ops time',
            'Standby', 'Down', 'Flat', 'Rough', 'Facilities', 'Dunes', 'Sabkha',
            'TCF', 'CTM', 'APP/CTM', 'PAD time',
        ]
        set_horizontal_cells(ws, 'A13', header, font_bold,
            Alignment(horizontal='center', wrap_text=True, vertical='top'))
        set_color(ws, 'A13:X13', color=lightblue)

        set_vertical_cells(ws, 'A14', dates, font_normal, Alignment())
        set_vertical_cells(ws, 'B14', prod_days, font_normal, Alignment())
        set_vertical_cells(ws, 'C14', vp_type['sp_t1'], font_normal, Alignment())
        set_vertical_cells(ws, 'D14', vp_type['sp_t2'], font_normal, Alignment())
        set_vertical_cells(ws, 'E14', vp_type['sp_t3'], font_normal, Alignment())
        set_vertical_cells(ws, 'F14', vp_type['sp_t4'], font_normal, Alignment())
        set_vertical_cells(ws, 'G14', vp_type['sp_t5'], font_normal, Alignment())
        set_vertical_cells(ws, 'H14', vp_type['skips'], font_normal, Alignment())
        set_vertical_cells(ws, 'I14', vp_type['total'], font_normal, Alignment())
        set_vertical_cells(ws, 'J14', vp_type['cumtotal'], font_normal, Alignment())
        set_vertical_cells(ws, 'K14', tm_type['rec_hours'], font_normal, Alignment())
        set_vertical_cells(ws, 'L14', tm_type['other_ops'], font_normal, Alignment())
        set_vertical_cells(ws, 'M14', tm_type['ops'], font_normal, Alignment())
        set_vertical_cells(ws, 'N14', tm_type['standby'], font_normal, Alignment())
        set_vertical_cells(ws, 'O14', tm_type['downtime'], font_normal, Alignment())
        set_vertical_cells(ws, 'P14', vp_dist['sp_t1'], font_normal, Alignment())
        set_vertical_cells(ws, 'Q14', vp_dist['sp_t2'], font_normal, Alignment())
        set_vertical_cells(ws, 'R14', vp_dist['sp_t3'], font_normal, Alignment())
        set_vertical_cells(ws, 'S14', vp_dist['sp_t4'], font_normal, Alignment())
        set_vertical_cells(ws, 'T14', vp_dist['sp_t5'], font_normal, Alignment())
        set_vertical_cells(ws, 'U14', vp_type['tcf'], font_normal, Alignment())
        set_vertical_cells(ws, 'V14', vp_type['ctm'], font_normal, Alignment())
        set_vertical_cells(ws, 'W14', vp_type['appctm'], font_normal, Alignment())
        set_vertical_cells(ws, 'X14', tm_type['padtime'], font_normal, Alignment())

        # and some formating of columns
        for c in ['K', 'L', 'M', 'N', 'O']:
            format_vertical(ws, f'{c}13:{c}100', '0.00')
        for c in ['P', 'Q', 'R', 'S', 'T', 'U', 'W', 'X']:
            format_vertical(ws, f'{c}13:{c}100', '0.0%')
        for c in ['I', 'J', 'V']:
            format_vertical(ws, f'{c}13:{c}100', '#,##0')

        # add row with the sums
        row_sum = [
            'Total', params['days'], vp_typesum['sp_t1'], vp_typesum['sp_t2'],
            vp_typesum['sp_t3'], vp_typesum['sp_t4'], vp_typesum['sp_t5'],
            vp_typesum['skips'], vp_typesum['total'], '', tm_typesum['rec_hours'],
            tm_typesum['other_ops'], tm_typesum['ops'], tm_typesum['standby'],
            tm_typesum['downtime'], vp_distsum['sp_t1'], vp_distsum['sp_t2'],
            vp_distsum['sp_t3'], vp_distsum['sp_t4'], vp_distsum['sp_t5'],
            vp_typesum['tcf'], vp_typesum['ctm'], vp_typesum['appctm'],
            tm_typesum['padtime'],
        ]
        set_horizontal_cells(ws, f'A{14+params["days"]}', row_sum, font_bold, Alignment())
        format_horizontal(ws, f'C{14+params["days"]}:H{14+params["days"]}', '#,##0')
        set_color(ws, f'A{14+params["days"]}:X{14+params["days"]}', color=lightblue)

    def create_tab_proj(self, proj_df):
        set_column_widths(self.wp, 'A', [12.5])
        proj_df['date'] = proj_df['date'].dt.date
        rows = dataframe_to_rows(proj_df, index=False, header=True)
        self.wp.freeze_panes = 'B2'

        for r_id, row in enumerate(rows, 1):
            for c_id, value in enumerate(row, 1):
                self.wp.cell(row=r_id, column=c_id, value=value)

    def create_mprreport(self):
        main_proj_df = self.collate_mpr_data(
            self.prod_series, self.time_series, self.hse_series
        )
        stype = None
        if self.ws_stypes:
            # there are more than one sourcetypes
            header_sourcetype = {
                'name': 'mixed source types',
                'sweep': '',
                'vibes': '',
                'moveup': '',
                'rechours': 24,
                'rate': self.prod_total['month_rate']
            }
            params = self.get_parameters(
                self.prod_total, self.times_total, header_sourcetype)
            self.create_tab_mpr(self.ws, params, main_proj_df)

            for stype_name in self.sourcetype_names:
                stype = self.day.project.sourcetypes.get(
                    sourcetype_name=stype_name
                )
                header_sourcetype = {
                    'name': stype_name,
                    'sweep': stype.mpr_sweep_length,
                    'vibes': stype.mpr_vibes,
                    'moveup': stype.mpr_moveup,
                    'rechours': stype.mpr_rec_hours,
                    'rate': ''
                }
                params = self.get_parameters(
                    self.prod_total, self.times_total, header_sourcetype)
                proj_df = self.collate_mpr_data(
                    self.prod_series_by_type[stype_name], self.time_series,
                    self.hse_series
                )
                self.create_tab_mpr(self.ws_stypes[stype_name], params, proj_df)

        else:
            # only a single sourcetype
            stype = self.day.project.sourcetypes.all()[0]
            header_sourcetype = {
                'name': stype.sourcetype_name,
                'sweep': stype.mpr_sweep_length,
                'vibes': stype.mpr_vibes,
                'moveup': stype.mpr_moveup,
                'rechours': stype.mpr_rec_hours,
                'rate': self.prod_total['month_rate']
            }
            params = self.get_parameters(
                self.prod_total, self.times_total, header_sourcetype)
            self.create_tab_mpr(self.ws, params, main_proj_df)

        self.create_tab_proj(main_proj_df)
        return save_excel(self.wb)
