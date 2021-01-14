''' writing excel with python
'''
import re
import io
from pathlib import Path
from openpyxl import Workbook, drawing
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.formatting.rule import CellIsRule
from daily_report.report_backend import ReportInterface
from daily_report.receiver_backend import ReceiverInterface
from seismicreport.vars import AVG_PERIOD, NO_DATE_STR


class ExcelReport:
    ''' class to create excel reports in CSR format
    '''
    def __init__(self, report_data, media_root, static_root):
        self.media_root = Path(media_root)
        self.static_root = Path(static_root)
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.sheet_view.showGridLines = False
        self.parse_data(report_data)

    def parse_data(self, report_data):
        self.report_date = report_data['report_date']
        self.project_table = report_data['project_table']
        self.daily_table = report_data['daily_table']
        self.recvr_table = report_data['receiver_table']
        self.csr_table = report_data['csr_table']
        self.proj_stats_table = report_data['proj_stats_table']
        self.block_stats_table = report_data['block_stats_table']
        self.hse_stats_table = report_data['hse_stats_table']
        self.csr_comment_table = report_data['csr_comment_table']

    def create_dailyreport(self):
        ''' method to create excel daily report
        '''
        img_width =310
        img_height = 150
        fontname = 'Tahoma'
        red = '00FF0000'
        green = '0000FF00'
        font_large_bold = Font(name=fontname, bold=True, size=11)
        font_normal = Font(name=fontname, size=9)
        font_bold = Font(name=fontname, bold=True, size=9)

        self.ws.column_dimensions['A'].width = 0.94
        self.ws.column_dimensions['B'].width = 0.75
        self.ws.column_dimensions['C'].width = 11.78
        self.ws.column_dimensions['D'].width = 10.89
        self.ws.column_dimensions['E'].width = 20.89
        self.ws.column_dimensions['F'].width = 0.56
        self.ws.column_dimensions['G'].width = 0.75
        self.ws.column_dimensions['H'].width = 14.11
        self.ws.column_dimensions['I'].width = 10.56
        self.ws.column_dimensions['J'].hidden = True
        self.ws.column_dimensions['K'].width = 13.22
        self.ws.column_dimensions['L'].width = 11.00

        # set logo
        img_logo = drawing.image.Image(self.static_root / 'img/display_icon.png')
        img_logo.width = 75
        img_logo.height = 75
        self.ws.add_image(img_logo, 'C4')

        # set title
        self.ws.merge_cells('B2:K2')
        self.ws['B2'].value = 'CSR DAILY REPORT'
        self.ws['B2'].alignment = Alignment(horizontal='center')
        self.ws['B2'].font = font_large_bold

        # set date
        self.ws.merge_cells('H3:I3')
        self.ws.merge_cells('K3:L3')
        self.set_vertical_cells(
            'H3', ['DATE'], font_large_bold, Alignment(horizontal='right'))
        self.set_vertical_cells('K3', [self.report_date], font_large_bold, Alignment())
        self.ws['K3'].font = Font(name=fontname, bold=True, size=11, color=red)

        # general project info
        self.set_vertical_cells(
            'D4', [key for key in self.project_table], font_bold, Alignment())
        self.set_vertical_cells(
            'E4', [val for _, val in self.project_table.items()], font_normal,
            Alignment(horizontal='right'))
        self.ws['E4'].font = font_bold
        self.ws['E5'].number_format = '#,##0'

        # daily stats
        self.set_vertical_cells(
            'H4', [key for key in self.daily_table], font_bold, Alignment())
        self.set_vertical_cells(
            'I4', [val for _, val in self.daily_table.items()], font_normal,
            Alignment(horizontal='right'))
        self.ws['I5'].number_format = '#,##0'
        self.ws['I6'].number_format = '#,##0'
        self.ws['I7'].number_format = '0.00%'
        self.ws['I8'].number_format = '0.00'
        self.ws['I9'].number_format = '0.00'
        self.ws['I10'].number_format = '0.00'
        self.ws['I11'].number_format = '#,##0'
        fill = PatternFill(bgColor=red)
        self.ws.conditional_formatting.add(
            'I7', CellIsRule(operator='lessThan', formula=[0.9], fill=fill))
        self.ws.conditional_formatting.add(
            'I8', CellIsRule(operator='lessThan', formula=[22], fill=fill))

        # receiver stats
        self.set_vertical_cells(
            'H12', [key for key in self.recvr_table
            if key != 'QC camp'], font_bold, Alignment())
        self.set_vertical_cells(
            'I12', [val for key, val in self.recvr_table.items()
            if key != 'QC camp'], font_normal, Alignment(horizontal='right'))
        self.ws['I12'].number_format = '#,##0'
        self.ws['I13'].number_format = '#,##0'
        self.ws['I14'].number_format = '0.00'
        self.ws['I15'].number_format = '#,##0'

        # XG01 and CSR
        self.set_vertical_cells(
            'D10', [key[:-3] for key in self.csr_table], font_bold,
            Alignment(horizontal='right'))
        self.set_vertical_cells(
            'E10', [val for _, val in self.csr_table.items()], font_normal,
            Alignment())

        # project stats
        self.ws.merge_cells('K4:L4')
        self.set_vertical_cells(
            'K4', ['Project Statistics'], font_large_bold,
            Alignment(horizontal='center'))
        self.set_vertical_cells(
            'K5', [key for key in self.proj_stats_table], font_bold, Alignment())
        self.set_vertical_cells(
            'L5', [val for _, val in self.proj_stats_table.items()], font_normal,
            Alignment(horizontal='right'))
        self.ws['L5'].number_format = '#,##0'
        self.ws['L6'].number_format = '0.00'
        self.ws['L7'].number_format = '#,##0'
        self.ws['L8'].number_format = '0.00%'

        # block stats
        self.ws.merge_cells('K11:L11')
        self.set_vertical_cells(
            'K11', ['Block Statistics'], font_large_bold,
            Alignment(horizontal='center'))
        self.set_vertical_cells(
            'K12', [key for key in self.block_stats_table], font_bold, Alignment())
        self.set_vertical_cells(
            'L12', [val for _, val in self.block_stats_table.items()], font_normal,
            Alignment(horizontal='right'))
        self.ws['L13'].number_format = '0.00'
        self.ws['L14'].number_format = '0.00%'

        # hse stats
        self.ws.merge_cells('K17:L17')
        self.set_vertical_cells(
            'K17', ['HSE Statistics'], font_large_bold,
            Alignment(horizontal='center'))
        self.set_vertical_cells(
            'K18', [key for key in self.hse_stats_table], font_bold, Alignment())
        self.set_vertical_cells(
            'L18', [val for _, val in self.hse_stats_table.items()], font_normal,
            Alignment(horizontal='right'))

        for i in range(18, 25):
            self.ws['L'+str(i)].number_format = '0;-0;;@'

        fill = PatternFill(bgColor=green)
        self.ws.conditional_formatting.add(
            'L24', CellIsRule(operator='greaterThan', formula=[9], fill=fill))

        # csr comment
        self.ws.merge_cells('B16:I26')
        self.set_vertical_cells(
            'B15', [key for key in self.csr_comment_table], font_bold, Alignment())
        self.set_vertical_cells(
            'B16', [val for _, val in self.csr_comment_table.items()], font_normal,
            Alignment(vertical='top', wrap_text=True))

        # add graphs
        img_daily_prod = drawing.image.Image(self.media_root / 'images/daily_prod.png')
        img_daily_prod.width = img_width
        img_daily_prod.height = img_height
        self.ws.add_image(img_daily_prod, 'C28')

        img_app_ctm = drawing.image.Image(self.media_root / 'images/app_ctm.png')
        img_app_ctm.width = img_width
        img_app_ctm.height = img_height
        self.ws.add_image(img_app_ctm, 'H28')

        img_cumul_prod = drawing.image.Image(self.media_root / 'images/cumul_prod.png')
        img_cumul_prod.width = img_width
        img_cumul_prod.height = img_height
        self.ws.add_image(img_cumul_prod, 'C37')

        img_rec_hours = drawing.image.Image(self.media_root / 'images/rec_hours.png')
        img_rec_hours.width = img_width
        img_rec_hours.height = img_height
        self.ws.add_image(img_rec_hours, 'H37')

        # set borders
        self.set_outer_border('B2:L46')
        self.set_outer_border('B2:L2')
        self.set_outer_border('B3:I15')
        self.set_outer_border('H4:I11')
        self.set_outer_border('H12:I15')
        self.set_outer_border('K4:L9')
        self.set_outer_border('K4:L4')
        self.set_outer_border('K10:L15')
        self.set_outer_border('K11:L11')
        self.set_outer_border('K17:L26')
        self.set_outer_border('K17:L17')
        self.set_outer_border('B16:I26')
        self.ws['I3'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'))

    def save_excel(self):
        ''' method to save excel data to a BytesIO buffer
        '''
        f_excel = io.BytesIO()
        self.wb.save(f_excel)
        f_excel.seek(0)
        return f_excel

    def set_vertical_cells(self, loc_init: str, values: list, font, alignment):
        # loc_init: format "A0" multiple capital letters followd by multiple numbers
        loc = re.match(r'^([A-Z]+)([0-9]+)$', loc_init)
        row_init = int(loc.group(2))
        col_init = loc.group(1)
        for i, value in enumerate(values):
            loc = col_init + str(row_init + i)
            cell = self.ws[loc]
            cell.value = value
            cell.font = font
            cell.alignment = alignment

    def set_outer_border(self, cell_range, style='thin'):
        cells =self.ws[cell_range]
        row = None
        for i, row in enumerate(cells):
            if i == 0:
                for col in row:
                    col.border = col.border + Border(top=Side(style=style))

            row[0].border = row[0].border + Border(left=Side(style=style))
            row[-1].border = row[-1].border + Border(right=Side(style=style))

        for col in row:
            col.border = col.border + Border(bottom=Side(style=style))


def collate_excel_dailyreport_data(day):
    rprt_iface = ReportInterface('')
    rcvr_iface = ReceiverInterface()
    totals_production, totals_time, totals_hse = rprt_iface.calc_totals(day)
    totals_receiver = rcvr_iface.calc_receiver_totals(day)
    # no need to call create_graphs as this has been done in DailyView

    project = day.project

    report_data = {}
    report_data['report_date'] = day.production_date.strftime('%#d %b %Y')

    if project.planned_start_date:
        proj_start_str = project.planned_start_date.strftime('%#d %b %Y')
        ops_days = (day.production_date - project.planned_start_date).days

    else:
        proj_start_str = NO_DATE_STR
        ops_days = NO_DATE_STR

    report_data['project_table'] = {
        'Project': project.project_name,
        'Project VPs': project.planned_vp,
        'Area (km\u00B2)': project.planned_area,
        'Proj. Start': proj_start_str,
        'Crew': project.crew_name,
    }

    report_data['daily_table'] = {
        'Oper Day': ops_days,
        'Total VPs': totals_production['total_sp'],
        'Target VPs': totals_production['ctm'][0],
        '% Target': totals_production['ctm'][1],
        'Recording Hrs': totals_time['rec_time'],
        'Standby Hrs': totals_time['standby'],
        'Downtime Hrs': totals_time['downtime'],
        'Skip VPs': totals_production['skips'],
    }

    report_data['receiver_table'] = {
        'Layout': totals_receiver['layout'],
        'Pickup': totals_receiver['pickup'],
        'QC field': totals_receiver['qc_field'],
        'QC camp': totals_receiver['qc_camp'],
        'Data upload': totals_receiver['upload'],
    }

    # cst table XG0 first then CSR
    xg0_staff = {
        f'{p.department}_{i:02}': p.name for i, p in
        enumerate(day.staff.filter(department__startswith='X').order_by('department'))
    }
    csr_staff = {
        f'{p.department}_{i:02}': p.name for i, p in
        enumerate(day.staff.filter(department__startswith='C').order_by('name'))
    }
    report_data['csr_table'] = {**xg0_staff, **csr_staff}

    # proj_stats
    proj_total = totals_production['proj_total']
    proj_skips = totals_production['proj_skips']
    if project.planned_vp > 0:
        proj_complete = (proj_total + proj_skips) / project.planned_vp

    else:
        proj_complete = 0

    report_data['proj_stats_table'] = {
        'Recorded VPs': proj_total,
        'Area (km\u00B2)': project.planned_area * proj_complete,
        'Skip VPs': proj_skips,
        '% Complete': proj_complete,
        'Est. Complete': rprt_iface.calc_est_completion_date(
            day, AVG_PERIOD, project.planned_vp, proj_complete),
    }

    # block_stats
    block = day.block
    totals_block = rprt_iface.calc_block_totals(day)
    if block and totals_block:
        block_name = block.block_name
        block_total = totals_block['block_total'] + totals_block['block_skips']
        block_complete = block_total / block.block_planned_vp
        block_area = block.block_planned_area * block_complete
        block_completion_date = rprt_iface.calc_est_completion_date(
            day, AVG_PERIOD, block.block_planned_vp, block_complete)

    else:
        block_name = ''
        block_complete = 0
        block_area = 0
        block_total = 0
        block_completion_date = NO_DATE_STR

    report_data['block_stats_table'] = {
        'Block': block_name,
        'Area (km\u00B2)': block_area,
        '% Complete': block_complete,
        'Est. Complete': block_completion_date,
    }

    report_data['hse_stats_table'] = {
        'LTI': totals_hse['lti'],
        'RWC': totals_hse['rwc'],
        'MTC': totals_hse['mtc'],
        'FAC': totals_hse['fac'],
        'NM/ Incidents': totals_hse['incident_nm'],
        'LSR': totals_hse['lsr_violations'],
        'STOP': totals_hse['stop'],
    }

    report_data['csr_comment_table'] = {
        'Comments': day.csr_comment,
    }

    return report_data
