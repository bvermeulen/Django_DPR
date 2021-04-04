''' module for creating excel monthly services '''
import calendar
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from seismicreport.utils.utils_excel import (
    set_vertical_cells, set_horizontal_cells, format_horizontal,
    set_outer_border, set_column_widths, save_excel,
)
from seismicreport.vars import CONTRACT

fontname = 'Tahoma'
font_large_bold = Font(name=fontname, bold=True, size=11)
font_normal = Font(name=fontname, size=9)
font_bold = Font(name=fontname, bold=True, size=9)
lightblue = 'D8EFF8'

#pylint: disable=line-too-long


class Mixin:

    @staticmethod
    def create_tab_services(ws, project, year, month):
        dayrange = range(1, calendar.monthrange(year, month)[1] + 1)
        days_in_month = len(dayrange)
        last_col = {31:'AG', 30: 'AF', 29: 'AE', 28: 'AD'}
        monthyear = f'{calendar.month_name[month]}, {year}'

        # set column widths
        set_column_widths(ws, 'A', [20.0])
        set_column_widths(ws, 'B', [4.0]*days_in_month + [8])

        # set the titles
        ws.merge_cells('P1:T1')
        ws['P1'].value = 'Monthly services'
        ws['P1'].font = font_large_bold

        for r in range(2, 6):
            ws.merge_cells(f'B{r}:E{r}')

        set_vertical_cells(ws, 'A2', ['Contract'], font_bold, Alignment(horizontal='left'))
        set_vertical_cells(ws, 'A3', ['Project'], font_bold, Alignment(horizontal='left'))
        set_vertical_cells(ws, 'A4', ['Crew'], font_bold, Alignment(horizontal='left'))
        set_vertical_cells(ws, 'A5', ['Month, Year'], font_bold, Alignment(horizontal='left'))

        set_vertical_cells(ws, 'B2', [CONTRACT], font_normal, Alignment(horizontal='right'))
        set_vertical_cells(ws, 'B3', [project.project_name], font_normal, Alignment(horizontal='right'))
        set_vertical_cells(ws, 'B4', [project.crew_name], font_normal, Alignment(horizontal='right'))
        set_vertical_cells(ws, 'B5', [monthyear], font_normal, Alignment(horizontal='right'))

        # set the header
        set_horizontal_cells(ws, 'A7', ['Item'], font_bold, Alignment())
        set_horizontal_cells(ws, 'B7', [*dayrange, 'Total'], font_bold, Alignment(horizontal='center'))
        set_outer_border(ws, f'A7:{last_col[days_in_month]}7')

        # set the services data
        r = 8
        for service in project.services.all():
            service_start_row = r
            set_horizontal_cells(ws, f'A{r}', [service.description], font_bold, Alignment())
            for task in service.tasks.all():
                r += 1
                set_horizontal_cells(ws, f'A{r}', [task.task_name], font_normal, Alignment())

                qties = {f'{d:02}': 0.0 for d in dayrange}
                task_qties = task.get_monthly_task_quantities(year, month)
                total = 0.0
                for dq in task_qties:
                    total += dq.quantity
                    qties[f'{dq.date.day:02}'] = dq.quantity
                row_values = list(qties.values()) + [total]
                set_horizontal_cells(ws, f'B{r}', row_values, font_normal, Alignment(horizontal='right'))
                format_horizontal(ws, f'B{r}:{last_col[days_in_month]}{r}', '0.0')

            set_outer_border(ws, f'A{service_start_row}:{last_col[days_in_month]}{r+1}')
            r += 2


class ExcelServiceReport(Mixin):

    def __init__(self, day):
        self.day = day
        self.wb = Workbook()
        self.wb.remove_sheet(self.wb.get_sheet_by_name('Sheet'))
        self.ws = self.wb.create_sheet('Services')
        self.ws.sheet_view.showGridLines = False

    def create_servicereport(self, year, month):
        self.create_tab_services(self.ws, self.day.project, year, month)
        return save_excel(self.wb)
