''' module to generate excel version of daily report
'''
from pathlib import Path
from openpyxl import Workbook, drawing
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.formatting.rule import CellIsRule
from daily_report.report_backend import ReportInterface
from seismicreport.utils.utils_excel import (
    set_vertical_cells, set_outer_border, save_excel
)
from seismicreport.vars import AVG_PERIOD, NO_DATE_STR, SS_2, IMG_SIZE


class ExcelDayReport:
    ''' class to create excel day report in CSR format
    '''
    def __init__(self, report_data, media_root, static_root):
        self.media_root = Path(media_root)
        self.static_root = Path(static_root)
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.sheet_view.showGridLines = False
        self.parse_data(report_data)

    def parse_data(self, report_data):
        self.report_date = report_data['report_date']
        self.project_table = report_data['project_table']
        self.daily_table = report_data['daily_table']
        self.recvr_table = report_data['receiver_table']
        self.csr_table = report_data['csr_table']
        self.proj_stats_table = report_data['proj_stats_table']
        self.block_stats_table = report_data['block_stats_table']
        self.hse_stats_table = report_data['hse_stats_table']
        self.csr_comment_table = report_data['csr_comment_table']

    def create_dailyreport(self):
        ''' method to create excel daily report
        '''
        fontname = 'Tahoma'
        red = '00FF0000'
        green = '0000FF00'
        font_large_bold = Font(name=fontname, bold=True, size=11)
        font_normal = Font(name=fontname, size=9)
        font_bold = Font(name=fontname, bold=True, size=9)

        self.ws.column_dimensions['A'].width = 0.94
        self.ws.column_dimensions['B'].width = 0.75
        self.ws.column_dimensions['C'].width = 11.78
        self.ws.column_dimensions['D'].width = 10.89
        self.ws.column_dimensions['E'].width = 20.89
        self.ws.column_dimensions['F'].width = 0.56
        self.ws.column_dimensions['G'].width = 0.75
        self.ws.column_dimensions['H'].width = 14.11
        self.ws.column_dimensions['I'].width = 10.56
        self.ws.column_dimensions['J'].hidden = True
        self.ws.column_dimensions['K'].width = 13.22
        self.ws.column_dimensions['L'].width = 11.00

        # set logo
        img_logo = drawing.image.Image(self.static_root / 'img/client_icon.png')
        img_logo.width = 75
        img_logo.height = 75
        self.ws.add_image(img_logo, 'C4')

        # set title
        self.ws.merge_cells('B2:K2')
        self.ws['B2'].value = 'CSR DAILY REPORT'
        self.ws['B2'].alignment = Alignment(horizontal='center')
        self.ws['B2'].font = font_large_bold

        # set date
        self.ws.merge_cells('H3:I3')
        self.ws.merge_cells('K3:L3')
        set_vertical_cells(
            self.ws, 'H3', ['DATE'], font_large_bold, Alignment(horizontal='right'))
        set_vertical_cells(
            self.ws, 'K3', [self.report_date], font_large_bold, Alignment())
        self.ws['K3'].font = Font(name=fontname, bold=True, size=11, color=red)

        # general project info
        set_vertical_cells(
            self.ws, 'D4', [key for key in self.project_table], font_bold, Alignment())
        set_vertical_cells(
            self.ws, 'E4', [val for _, val in self.project_table.items()], font_normal,
            Alignment(horizontal='right'))
        self.ws['E4'].font = font_bold
        self.ws['E5'].number_format = '#,##0'

        # daily stats
        set_vertical_cells(
            self.ws, 'H4', [key for key in self.daily_table], font_bold, Alignment())
        set_vertical_cells(
            self.ws, 'I4', [val for _, val in self.daily_table.items()], font_normal,
            Alignment(horizontal='right'))
        self.ws['I5'].number_format = '#,##0'
        self.ws['I6'].number_format = '#,##0'
        self.ws['I7'].number_format = '0.00%'
        self.ws['I8'].number_format = '0.00'
        self.ws['I9'].number_format = '0.00'
        self.ws['I10'].number_format = '0.00'
        self.ws['I11'].number_format = '0.00'
        self.ws['I12'].number_format = '#,##0'
        fill = PatternFill(bgColor=red)
        self.ws.conditional_formatting.add(
            'I7', CellIsRule(operator='lessThan', formula=[0.9], fill=fill))
        self.ws.conditional_formatting.add(
            'I8', CellIsRule(operator='lessThan', formula=[22], fill=fill))

        # receiver stats
        not_shown = ['Node charged', 'Node repair']
        set_vertical_cells(
            self.ws, 'H13', [key for key in self.recvr_table
            if key not in not_shown], font_bold, Alignment())
        set_vertical_cells(
            self.ws, 'I13', [val for key, val in self.recvr_table.items()
            if key not in not_shown], font_normal, Alignment(horizontal='right'))
        self.ws['I13'].number_format = '#,##0'
        self.ws['I14'].number_format = '#,##0'
        self.ws['I15'].number_format = '#,##0'
        self.ws['I16'].number_format = '#,##0'

        # XG01 and CSR
        set_vertical_cells(
            self.ws, 'D10', [key[:-3] for key in self.csr_table], font_bold,
            Alignment(horizontal='right'))
        set_vertical_cells(
            self.ws, 'E10', [val for _, val in self.csr_table.items()], font_normal,
            Alignment())

        # project stats
        self.ws.merge_cells('K4:L4')
        set_vertical_cells(
            self.ws, 'K4', ['Project Statistics'], font_large_bold,
            Alignment(horizontal='center'))
        set_vertical_cells(
            self.ws, 'K5', [key for key in self.proj_stats_table], font_bold,
            Alignment())
        set_vertical_cells(
            self.ws, 'L5', [val for _, val in self.proj_stats_table.items()], font_normal,
            Alignment(horizontal='right'))
        self.ws['L5'].number_format = '#,##0'
        self.ws['L6'].number_format = '0.00'
        self.ws['L7'].number_format = '#,##0'
        self.ws['L8'].number_format = '0.00%'

        # block stats
        self.ws.merge_cells('K10:L10')
        set_vertical_cells(
            self.ws, 'K10', ['Block Statistics'], font_large_bold,
            Alignment(horizontal='center'))
        set_vertical_cells(
            self.ws, 'K11', [key for key in self.block_stats_table], font_bold,
            Alignment())
        set_vertical_cells(
            self.ws, 'L11', [val for _, val in self.block_stats_table.items()],
            font_normal, Alignment(horizontal='right'))
        self.ws['L12'].number_format = '0.00'
        self.ws['L13'].number_format = '0.00%'

        # hse stats
        self.ws.merge_cells('K15:L15')
        set_vertical_cells(
            self.ws, 'K15', ['HSE Statistics'], font_large_bold,
            Alignment(horizontal='center'))
        set_vertical_cells(
            self.ws, 'K16', [key for key in self.hse_stats_table], font_bold, Alignment())
        set_vertical_cells(
            self.ws, 'L16', [val for _, val in self.hse_stats_table.items()], font_normal,
            Alignment(horizontal='right'))

        for i in range(16, 23):
            self.ws['L'+str(i)].number_format = '0;-0;;@'

        fill = PatternFill(bgColor=green)
        self.ws.conditional_formatting.add(
            'L23', CellIsRule(operator='greaterThan', formula=[9], fill=fill))

        # csr comment
        self.ws.merge_cells('B17:I27')
        set_vertical_cells(
            self.ws, 'B16', [key for key in self.csr_comment_table], font_bold,
            Alignment())
        set_vertical_cells(
            self.ws, 'B17', [val for _, val in self.csr_comment_table.items()],
            font_normal, Alignment(vertical='top', wrap_text=True))

        # add graphs
        width, height = IMG_SIZE
        img_daily_prod = drawing.image.Image(self.media_root / 'images/daily_prod.png')
        img_daily_prod.width = width
        img_daily_prod.height = height
        self.ws.add_image(img_daily_prod, 'C29')

        img_cumul_prod = drawing.image.Image(self.media_root / 'images/cumul_prod.png')
        img_cumul_prod.width = width
        img_cumul_prod.height = height
        self.ws.add_image(img_cumul_prod, 'H29')

        img_rec_hours = drawing.image.Image(self.media_root / 'images/rec_hours.png')
        img_rec_hours.width = width
        img_rec_hours.height = height
        self.ws.add_image(img_rec_hours, 'C42')

        img_app_ctm = drawing.image.Image(self.media_root / 'images/app_ctm_ratio.png')
        img_app_ctm.width = width
        img_app_ctm.height = height
        self.ws.add_image(img_app_ctm, 'H42')

        # set borders
        set_outer_border(self.ws, 'B2:L55')
        set_outer_border(self.ws, 'B2:L2')
        set_outer_border(self.ws, 'B3:I16')
        set_outer_border(self.ws, 'H4:I12')
        set_outer_border(self.ws, 'H13:I16')
        set_outer_border(self.ws, 'K4:L9')
        set_outer_border(self.ws, 'K4:L4')
        set_outer_border(self.ws, 'K10:L14')
        set_outer_border(self.ws, 'K10:L10')
        set_outer_border(self.ws, 'K15:L27')
        set_outer_border(self.ws, 'K15:L15')
        set_outer_border(self.ws, 'B17:I27')
        self.ws['I3'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'))

        return save_excel(self.wb)


def collate_excel_dailyreport_data(day):
    r_iface = ReportInterface('')
    (
        _,
        totals_production,
        totals_time,
        totals_receiver,
        totals_hse
     ) = r_iface.calc_totals(day)
    # no need to call create_daily_graphs as this has been done in DailyView

    project = day.project

    report_data = {}
    report_data['report_date'] = day.production_date.strftime('%#d %b %Y')

    if project.start_report:
        ops_days = (day.production_date - project.start_report).days + 1

    else:
        ops_days = NO_DATE_STR

    if project.planned_start_date:
        proj_start_str = project.planned_start_date.strftime('%#d %b %Y')

    else:
        proj_start_str = NO_DATE_STR

    report_data['project_table'] = {
        'Project': project.project_name,
        'Project VPs': project.planned_vp,
        f'Area (km{SS_2})': project.planned_area,
        'Proj. Start': proj_start_str,
        'Crew': project.crew_name,
    }

    report_data['daily_table'] = {
        'Oper Day': ops_days,
        'Total VPs': totals_production['day_total'],
        'Target VPs': totals_production['day_ctm'],
        '% Target': totals_production['day_appctm'],
        'Recording Hrs': totals_time['day_rec_time'],
        'Ops Hrs': totals_time['day_ops_time'],
        'Standby Hrs': totals_time['day_standby'],
        'Downtime Hrs': totals_time['day_downtime'],
        'Skip VPs': totals_production['day_skips'],
    }

    report_data['receiver_table'] = {
        'Layout': totals_receiver['day_layout'],
        'Pickup': totals_receiver['day_pickup'],
        'Node download': totals_receiver['day_node_download'],
        'Node charged': totals_receiver['day_node_charged'],
        'Node failure': totals_receiver['day_node_failure'],
        'Node repair': totals_receiver['day_node_repair'],
    }

    # cst table XG0 first then CSR
    xg0_staff = {
        f'{p.department}_{i:02}': p.name for i, p in
        enumerate(day.staff.filter(department__startswith='X').order_by('department'))
    }
    csr_staff = {
        f'{p.department}_{i:02}': p.name for i, p in
        enumerate(day.staff.filter(department__startswith='C').order_by('name'))
    }
    report_data['csr_table'] = {**xg0_staff, **csr_staff}

    # proj_stats
    proj_total = totals_production['proj_total']
    proj_skips = totals_production['proj_skips']
    if project.planned_vp > 0:
        proj_complete = (proj_total + proj_skips) / project.planned_vp

    else:
        proj_complete = 0

    report_data['proj_stats_table'] = {
        'Recorded VPs': proj_total,
        f'Area (km{SS_2})': project.planned_area * proj_complete,
        'Skip VPs': proj_skips,
        '% Complete': proj_complete,
        'Est. Complete': r_iface.calc_est_completion_date(
            day, AVG_PERIOD, project.planned_vp, proj_complete),
    }

    # block_stats
    block = day.block
    totals_block = r_iface.calc_block_totals(day)
    if block and totals_block:
        block_name = block.block_name
        block_total = totals_block['block_total'] + totals_block['block_skips']
        block_complete = block_total / block.block_planned_vp
        block_area = block.block_planned_area * block_complete
        block_completion_date = r_iface.calc_est_completion_date(
            day, AVG_PERIOD, block.block_planned_vp, block_complete)

    else:
        block_name = ''
        block_complete = 0
        block_area = 0
        block_total = 0
        block_completion_date = NO_DATE_STR

    report_data['block_stats_table'] = {
        'Block': block_name,
        f'Area (km{SS_2})': block_area,
        '% Complete': block_complete,
        'Est. Complete': block_completion_date,
    }

    report_data['hse_stats_table'] = {
        'LTI': totals_hse['day_lti'],
        'RWC': totals_hse['day_rwc'],
        'MTC': totals_hse['day_mtc'],
        'FAC': totals_hse['day_fac'],
        'NM/ Incidents': totals_hse['day_incident_nm'],
        'LSR': totals_hse['day_lsr_violations'],
        'Inspections': totals_hse['day_audits'],
        'STOP': totals_hse['day_stop'],
    }

    report_data['csr_comment_table'] = {
        'Comment': day.csr_comment
    }

    return report_data
