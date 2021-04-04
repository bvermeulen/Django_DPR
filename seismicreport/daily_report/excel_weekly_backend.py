''' module to generate excel version of weekly report
'''
from pathlib import Path
import numpy as np
from openpyxl import Workbook, drawing
from openpyxl.styles import Alignment, Font
from daily_report.report_backend import ReportInterface
from daily_report.week_backend import WeekInterface
import daily_report.graph_backend as _graph_backend
from seismicreport.utils.utils_excel import (
    set_vertical_cells, set_horizontal_cells, format_vertical, format_horizontal,
    conditional_format, set_column_widths, set_border, set_outer_border,
    set_color, save_excel, get_row_column,
)
from seismicreport.utils.utils_funcs import nan_array
from seismicreport.vars import (
    STOP_TARGET, PROD_TARGET, REC_TARGET, AVG_PERIOD, NO_DATE_STR, SS_2, IMG_SIZE,
    source_prod_schema,
)

fontname = 'Tahoma'
font_large_bold = Font(name=fontname, bold=True, size=11)
font_normal = Font(name=fontname, size=9)
font_bold = Font(name=fontname, bold=True, size=9)
red = '00FF0000'
green ='0000FF00'
orange = 'FFA500'
lightblue = 'D8EFF8'
float_hide_zero = '0.00;-0;;@'
int_hide_zero = '#,##0;0;;@'


class ExcelWeekReport(_graph_backend.Mixin):
    ''' class to create excel week report in CSR format
    '''
    def __init__(self, report_data, media_dir, static_dir):
        self.media_dir = Path(media_dir)
        self.static_dir = Path(static_dir)
        self.wb = Workbook()
        self.wb.remove_sheet(self.wb.get_sheet_by_name('Sheet'))
        self.wsw = self.wb.create_sheet('CSR_weekly')
        self.wst = self.wb.create_sheet('Times')
        self.wsp = self.wb.create_sheet('Production')
        self.wsg = self.wb.create_sheet('Graphs')
        self.wsw.sheet_view.showGridLines = False
        self.wst.sheet_view.showGridLines = False
        self.wsp.sheet_view.showGridLines = False
        self.wsg.sheet_view.showGridLines = False

        self.parse_data(report_data)

    def parse_data(self, report_data):
        self.report_date = report_data['report_date']
        self.proj_days = report_data['proj_days']
        self.month_days = report_data['month_days']
        self.author = report_data['author_table']
        self.comment = report_data['comment_table']
        self.project = report_data['project_table']
        self.proj_stats = report_data['proj_stats_table']
        self.hse_stats = list(zip(
            report_data['hse_stats_week_table'].keys(),
            report_data['hse_stats_week_table'].values(),
            report_data['hse_stats_month_table'].values(),
            report_data['hse_stats_proj_table'].values()
        ))
        self.prod_stats = list(zip(
            report_data['prod_stats_week_table'].keys(),
            report_data['prod_stats_week_table'].values(),
            report_data['prod_stats_month_table'].values(),
            report_data['prod_stats_proj_table'].values()
        ))
        self.days_times = report_data['days_times']
        self.weeks_times = report_data['weeks_times']
        self.days_prod =  report_data['days_prod']
        self.weeks_prod = report_data['weeks_prod']

        self.week_terrain = report_data['week_terrain']
        self.proj_terrain = report_data['proj_terrain']

        self.prod_series = report_data['prod_series']
        self.time_series = report_data['time_series']


    def create_tab_weekly(self):
        ''' method to create excel weekly report main tab
        '''
        set_column_widths(self.wsw, 'A',
            [0.94, 0.75, 12.56, 10.22, 19.89, 4.11, 0.81,
            11.67, 11.22, 12.56, 11.22, 0.94]
        )

        # set title
        self.wsw.merge_cells('B2:L2')
        self.wsw['B2'].value = 'CSR WEEKLY REPORT'
        self.wsw['B2'].alignment = Alignment(horizontal='center')
        self.wsw['B2'].font = font_large_bold

        # set logo
        img_logo = drawing.image.Image(self.static_dir / 'img/client_icon.png')
        img_logo.width = 75
        img_logo.height = 75
        self.wsw.add_image(img_logo, 'C4')

        # set project general
        set_vertical_cells(self.wsw, 'D4', [k for k in self.project], font_bold,
            Alignment())
        set_vertical_cells(self.wsw, 'E4', [v for v in self.project.values()],
            font_normal, Alignment(horizontal='right'))
        self.wsw['E5'].number_format = '#,##0'
        self.wsw['E6'].number_format = '0.00'

        # set author
        self.wsw.merge_cells('H4:I4')
        self.wsw.merge_cells('H5:I5')
        set_vertical_cells(self.wsw, 'H4', [k for k in self.author], font_bold,
            Alignment(horizontal='center'))
        set_vertical_cells(self.wsw, 'H5', [v for v in self.author.values()],
            font_normal, Alignment(horizontal='center'))

        # set date
        self.wsw.merge_cells('H3:I3')
        self.wsw.merge_cells('J3:K3')
        set_vertical_cells(
            self.wsw, 'H3', ['DATE'], font_large_bold, Alignment(horizontal='right'))
        set_vertical_cells(
            self.wsw, 'J3', [self.report_date], font_large_bold, Alignment())
        self.wsw['J3'].font = Font(name=fontname, bold=True, size=11, color=red)

        # set comments
        self.wsw.merge_cells('B11:F11')
        self.wsw.merge_cells('B12:F35')
        set_vertical_cells(self.wsw, 'B11', [k for k in self.comment], font_bold,
            Alignment())
        set_vertical_cells(self.wsw, 'B12', [v for v in self.comment.values()],
            font_normal, Alignment(vertical='top', wrap_text=True))

        # set project statistics
        self.wsw.merge_cells('J4:K4')
        set_vertical_cells(self.wsw, 'J4', ['Project Statistics'], font_bold,
            Alignment(horizontal='center'))
        set_vertical_cells(self.wsw, 'J5', [k for k in self.proj_stats], font_bold,
            Alignment())
        set_vertical_cells(self.wsw, 'K5', [v for v in self.proj_stats.values()],
            font_normal, Alignment(horizontal='right'))
        self.wsw['K5'].number_format = '#,##0'
        self.wsw['K6'].number_format = '0.00'
        self.wsw['K7'].number_format = '#,##0'
        self.wsw['K8'].number_format = '0.00%'

        # set hse statistic
        self.wsw.merge_cells('H11:K11')
        set_vertical_cells(self.wsw, 'H11', ['HSE Statistics'], font_bold,
            Alignment(horizontal='center'))
        set_horizontal_cells(self.wsw, 'I12', ['Week', 'Month', 'Project'], font_bold,
            Alignment(horizontal='center'))
        set_vertical_cells(self.wsw, 'H13', [v[0] for v in self.hse_stats], font_normal,
            Alignment())
        for row, item in enumerate(self.hse_stats):
            set_horizontal_cells(self.wsw, f'I{row+13}', [item[1], item[2], item[3]],
                font_normal, Alignment(horizontal='right'))

        f_vals = np.array(STOP_TARGET) * 7
        conditional_format(self.wsw, 'I19', f_vals, (red, orange, green))
        f_vals = np.array(STOP_TARGET) * self.month_days
        conditional_format(self.wsw, 'J19', f_vals, (red, orange, green))
        f_vals = np.array(STOP_TARGET) * self.proj_days
        conditional_format(self.wsw, 'K19', f_vals, (red, orange, green))

        # set production statistic
        self.wsw.merge_cells('H24:K24')
        set_vertical_cells(self.wsw, 'H24', ['Production Statistics'], font_bold,
            Alignment(horizontal='center'))
        set_horizontal_cells(self.wsw, 'I25', ['Week', 'Month', 'Project'], font_bold,
            Alignment(horizontal='center'))
        set_vertical_cells(self.wsw, 'H26', [v[0] for v in self.prod_stats], font_normal,
            Alignment())
        for row, item in enumerate(self.prod_stats):
            set_horizontal_cells(self.wsw, f'I{row+26}', [item[1], item[2], item[3]],
                font_normal, Alignment(horizontal='right'))

        format_horizontal(self.wsw, 'I26:K26', '#,##0')
        format_horizontal(self.wsw, 'I27:K27', '#,##0')
        format_horizontal(self.wsw, 'I28:K28', '0.00%')
        format_horizontal(self.wsw, 'I29:K29', '#,##0')
        format_horizontal(self.wsw, 'I30:K30', '0.00')
        format_horizontal(self.wsw, 'I31:K31', '#,##0')
        format_horizontal(self.wsw, 'I32:K32', '0.00')
        format_horizontal(self.wsw, 'I33:K33', '#,##0')
        format_horizontal(self.wsw, 'I34:K34', '0.00%')

        f_vals = np.array(PROD_TARGET) * 1
        conditional_format(self.wsw, 'I28', f_vals, (red, orange, green))
        conditional_format(self.wsw, 'J28', f_vals, (red, orange, green))
        conditional_format(self.wsw, 'K28', f_vals, (red, orange, green))

        f_vals = np.array(REC_TARGET) * 7
        conditional_format(self.wsw, 'I29', f_vals, (red, orange, green))
        f_vals = np.array(REC_TARGET) * self.month_days
        conditional_format(self.wsw, 'J29', f_vals, (red, orange, green))
        f_vals = np.array(REC_TARGET) * self.proj_days
        conditional_format(self.wsw, 'K29', f_vals, (red, orange, green))

        # add graphs
        width, height = IMG_SIZE

        img_daily_prod = drawing.image.Image(
            self.media_dir / 'images/cumul_app_ctm.png')
        img_daily_prod.width = width
        img_daily_prod.height = height
        self.wsw.add_image(img_daily_prod, 'C37')

        img_daily_prod = drawing.image.Image(
            self.media_dir / 'images/pie_proj_terrain.png')
        img_daily_prod.width = width
        img_daily_prod.height = height
        self.wsw.add_image(img_daily_prod, 'H37')

        img_daily_prod = drawing.image.Image(
            self.media_dir / 'images/rec_hours.png')
        img_daily_prod.width = width
        img_daily_prod.height = height
        self.wsw.add_image(img_daily_prod, 'C51')

        img_daily_prod = drawing.image.Image(
            self.media_dir / 'images/app_ctm_ratio.png')
        img_daily_prod.width = width
        img_daily_prod.height = height
        self.wsw.add_image(img_daily_prod, 'H51')

        # set borders
        set_outer_border(self.wsw, 'B2:L64')
        set_outer_border(self.wsw, 'B2:L2')
        set_outer_border(self.wsw, 'J4:L4')
        set_outer_border(self.wsw, 'J5:L9')
        set_outer_border(self.wsw, 'B11:F11')
        set_outer_border(self.wsw, 'B12:F35')
        set_outer_border(self.wsw, 'H11:L11')
        set_outer_border(self.wsw, 'H12:L22')
        set_outer_border(self.wsw, 'H24:L24')
        set_outer_border(self.wsw, 'H25:L35')

    def create_tab_times(self):
        ''' method to create excel weekly tab for times
        '''
        set_column_widths(self.wst, 'A',
            [0.94, 21.56, 9.22, 8.33, 8.33, 8.33, 8.33, 8.33,
            10.11, 12.33, 9.89, 8.33,
            8.33, 8.33, 8.33, 8.33, 8.33, 8.33, 8.33, 8.33,
            8.33, 8.33, 9.56,
            ]
        )
        # set day times
        self.wst.merge_cells('B1:V1')
        set_vertical_cells(self.wst, 'B1', ['Daily times'], font_bold,
            Alignment(horizontal='center'))
        self.wst.merge_cells('C2:H2')
        set_vertical_cells(self.wst, 'C2', ['Operational time'], font_bold,
            Alignment(horizontal='center'))
        self.wst.merge_cells('I2:L2')
        set_vertical_cells(self.wst, 'I2', ['Standby time'], font_bold,
            Alignment(horizontal='center'))
        self.wst.merge_cells('M2:S2')
        set_vertical_cells(self.wst, 'M2', ['Downtime'], font_bold,
            Alignment(horizontal='center'))
        self.wst.merge_cells('T2:V2')
        set_vertical_cells(self.wst, 'T2', ['Totals'], font_bold,
            Alignment(horizontal='center'))

        set_horizontal_cells(self.wst, 'B3', self.days_times['header'], font_bold,
            Alignment(horizontal='center', vertical='top', wrap_text=True,))

        row, col = get_row_column('B4')
        weeks_total = np.zeros(len(self.days_times[0]) - 1)
        for key in range(0, 7):
            vals = self.days_times[key]
            loc = col + str(row + key)
            set_horizontal_cells(self.wst, loc, vals, font_normal,
                Alignment(horizontal='right'))

            self.wst[f'B{row + key}'].alignment = Alignment(horizontal='center')
            format_range = f'C{row + key}:V{row + key}'
            format_horizontal(self.wst, format_range, float_hide_zero)

            weeks_total += nan_array(vals[1:]).astype(np.float)

        set_vertical_cells(self.wst, 'B11', ['Weeks total'], font_bold,
            Alignment(horizontal='center'))
        set_horizontal_cells(self.wst, 'C11', weeks_total, font_bold,
            Alignment(horizontal='right'))
        format_horizontal(self.wst, 'C11:V11', '0.00')

        # set borders day times
        set_outer_border(self.wst, 'C2:H2')
        set_outer_border(self.wst, 'I2:L2')
        set_outer_border(self.wst, 'M2:S2')
        set_outer_border(self.wst, 'T2:V2')
        set_border(self.wst, 'B3:v11')

        # set week times
        self.wst.merge_cells('B13:V13')
        set_vertical_cells(self.wst, 'B13', ['Weekly times'], font_bold,
            Alignment(horizontal='center'))
        self.wst.merge_cells('C14:H14')
        set_vertical_cells(self.wst, 'C14', ['Operational time'], font_bold,
            Alignment(horizontal='center'))
        self.wst.merge_cells('I14:L14')
        set_vertical_cells(self.wst, 'I14', ['Standby time'], font_bold,
            Alignment(horizontal='center'))
        self.wst.merge_cells('M14:S14')
        set_vertical_cells(self.wst, 'M14', ['Downtime'], font_bold,
            Alignment(horizontal='center'))
        self.wst.merge_cells('T14:V14')
        set_vertical_cells(self.wst, 'T14', ['Totals'], font_bold,
            Alignment(horizontal='center'))

        set_horizontal_cells(self.wst, 'B15', self.weeks_times['header'], font_bold,
            Alignment(horizontal='center', vertical='top', wrap_text=True,))

        row, col = get_row_column('B16')
        for key in range(0, 6):
            vals = self.weeks_times[key]
            loc = col + str(row + key)
            set_horizontal_cells(self.wst, loc, vals, font_normal,
                Alignment(horizontal='right'))

            self.wst[f'B{row + key}'].alignment = Alignment(horizontal='center')
            format_range = f'C{row + key}:V{row + key}'
            format_horizontal(self.wst, format_range, float_hide_zero)

        # set borders week times
        set_outer_border(self.wst, 'C14:H14')
        set_outer_border(self.wst, 'I14:L14')
        set_outer_border(self.wst, 'M14:S14')
        set_outer_border(self.wst, 'T14:V14')
        set_border(self.wst, 'B15:v21')
        set_color(self.wst, 'B21:V21', color=lightblue)

    def create_tab_production(self):
        ''' method to create excel weekly tab for production
        '''
        set_column_widths(self.wsp, 'A',
            [0.94, 23.50, 12.33, 12.33, 12.33, 12.33, 12.33, 12.33, 12.33, 12.33, 12.33,
            ]
        )
        # set day production
        self.wsp.merge_cells('B1:K1')
        set_vertical_cells(self.wsp, 'B1', ['Daily production'], font_bold,
            Alignment(horizontal='center'))

        set_horizontal_cells(self.wsp, 'B2', self.days_prod['header'], font_bold,
            Alignment(horizontal='center', vertical='top', wrap_text=True,))

        row, col = get_row_column('B3')
        weeks_total = np.zeros(len(self.days_prod[0])-1)
        for key in range(0, 7):
            vals = self.days_prod[key]
            loc = col + str(row + key)
            set_horizontal_cells(self.wsp, loc, vals, font_normal,
                Alignment(horizontal='right'))

            self.wsp[f'B{row + key}'].alignment = Alignment(horizontal='center')
            format_range = f'C{row + key}:C{row + key}'
            format_horizontal(self.wsp, format_range, int_hide_zero)
            format_range = f'D{row + key}:D{row + key}'
            format_horizontal(self.wsp, format_range, float_hide_zero)
            format_range = f'E{row + key}:K{row + key}'
            format_horizontal(self.wsp, format_range, int_hide_zero)

            weeks_total += nan_array(vals[1:]).astype(np.float)

        # skip the 3rd and 6 and 11th elements
        weeks_total = [*weeks_total[0:2], '', *weeks_total[3:10]]
        set_vertical_cells(self.wsp, 'B10', ['Weeks total'], font_bold,
            Alignment(horizontal='center'))
        set_horizontal_cells(self.wsp, 'C10', weeks_total, font_bold,
            Alignment(horizontal='right'))
        format_horizontal(self.wsp, 'C10:C10', '#,##0')
        format_horizontal(self.wsp, 'D10:D10', '0.00')
        format_horizontal(self.wsp, 'E10:K10', '#,##0')

        # set borders
        set_border(self.wsp, 'B2:M10')

        # set week production
        self.wsp.merge_cells('B13:K13')
        set_vertical_cells(self.wsp, 'B13', ['Weekly production'], font_bold,
            Alignment(horizontal='center'))

        set_horizontal_cells(self.wsp, 'B14', self.weeks_prod['header'], font_bold,
            Alignment(horizontal='center', vertical='top', wrap_text=True,))

        row, col = get_row_column('B15')
        for key in range(0, 6):
            vals = self.weeks_prod[key]
            loc = col + str(row + key)
            set_horizontal_cells(self.wsp, loc, vals, font_normal,
                Alignment(horizontal='right'))

            self.wsp[f'B{row + key}'].alignment = Alignment(horizontal='center')
            format_range = f'C{row + key}:C{row + key}'
            format_horizontal(self.wsp, format_range, int_hide_zero)
            format_range = f'D{row + key}:D{row + key}'
            format_horizontal(self.wsp, format_range, float_hide_zero)
            format_range = f'E{row + key}:K{row + key}'
            format_horizontal(self.wsp, format_range, int_hide_zero)

        # set borders
        set_border(self.wsp, 'B14:M20')
        set_color(self.wsp, 'B20:M20', color=lightblue)


        # set terrain types for week
        self.wsp.merge_cells('C22:E22')
        set_vertical_cells(self.wsp, 'C22', ['Week terrain'], font_bold,
            Alignment(horizontal='center'))
        set_horizontal_cells(self.wsp, 'D23', ['VPs', 'Percentage'], font_bold,
            Alignment(horizontal='center'))
        set_vertical_cells(self.wsp, 'C24', [key for key in self.week_terrain],
            font_normal, Alignment(horizontal='left'))

        vals = []
        percs = []
        for val in self.week_terrain.values():
            vals.append(val)
            if self.week_terrain['total'] > 0:
                perc = val / self.week_terrain['total']

            else:
                perc = np.nan

            percs.append(perc)

        set_vertical_cells(self.wsp, 'D24', vals, font_normal,
            Alignment(horizontal='right'))
        set_vertical_cells(self.wsp, 'E24', percs, font_normal,
            Alignment(horizontal='right'))
        format_vertical(self.wsp, 'D24:D30', '#,##0')
        format_vertical(self.wsp, 'E24:E30', '0.00%')
        set_border(self.wsp, 'D23:E23')
        set_border(self.wsp, 'C24:E30')

        # set terrain types for project
        self.wsp.merge_cells('G22:I22')
        set_vertical_cells(self.wsp, 'G22', ['Project terrain'], font_bold,
            Alignment(horizontal='center'))
        set_horizontal_cells(self.wsp, 'H23', ['VPs', 'Percentage'], font_bold,
            Alignment(horizontal='center'))
        set_vertical_cells(self.wsp, 'G24', [key for key in self.proj_terrain],
            font_normal, Alignment(horizontal='left'))

        vals = []
        percs = []
        for val in self.proj_terrain.values():
            vals.append(val)
            if self.proj_terrain['total'] > 0:
                perc = val / self.proj_terrain['total']

            else:
                perc = np.nan

            percs.append(perc)

        set_vertical_cells(self.wsp, 'H24', vals, font_normal,
            Alignment(horizontal='right'))
        set_vertical_cells(self.wsp, 'I24', percs, font_normal,
            Alignment(horizontal='right'))
        format_vertical(self.wsp, 'H24:H30', '#,##0')
        format_vertical(self.wsp, 'I24:I30', '0.00%')
        set_border(self.wsp, 'H23:I23')
        set_border(self.wsp, 'G24:I30')

    def create_tab_graphs(self):
        width, height = IMG_SIZE

        img_daily_prod = drawing.image.Image(
            self.media_dir / 'images/bar_week_production.png')
        img_daily_prod.width = width
        img_daily_prod.height = height
        self.wsg.add_image(img_daily_prod, 'B2')

        img_daily_prod = drawing.image.Image(
            self.media_dir / 'images/pie_week_terrain.png')
        img_daily_prod.width = width
        img_daily_prod.height = height
        self.wsg.add_image(img_daily_prod, 'H2')

        img_daily_prod = drawing.image.Image(
            self.media_dir / 'images/pie_week_times.png')
        img_daily_prod.width = width
        img_daily_prod.height = height
        self.wsg.add_image(img_daily_prod, 'B16')

        img_daily_prod = drawing.image.Image(
            self.media_dir / 'images/bar_day_production.png')
        img_daily_prod.width = width
        img_daily_prod.height = height
        self.wsg.add_image(img_daily_prod, 'H16')

        img_daily_prod = drawing.image.Image(
            self.media_dir / 'images/bar_day_rechours.png')
        img_daily_prod.width = width
        img_daily_prod.height = height
        self.wsg.add_image(img_daily_prod, 'B30')

        img_daily_prod = drawing.image.Image(
            self.media_dir / 'images/bar_day_vphr.png')
        img_daily_prod.width = width
        img_daily_prod.height = height
        self.wsg.add_image(img_daily_prod, 'H30')

    def create_weekreport(self):
        self.create_tab_weekly()
        self.create_tab_times()
        self.create_tab_production()
        self.create_weekly_graphs()
        self.create_tab_graphs()
        return save_excel(self.wb)


def collate_excel_weekreport_data(day):
    r_iface = ReportInterface('')
    w_iface = WeekInterface('')
    project = day.project

    report_data = {}
    _, totals_prod, totals_time, _, totals_hse = r_iface.calc_totals(day)
    days, weeks = w_iface.collate_weekdata(day)
    (
        _, report_data['prod_series'], report_data['time_series'],
        _, _
    ) = r_iface.series

    report_data['report_date'] = day.production_date.strftime('%#d %b %Y')
    report_data['month_days'] = day.production_date.day
    if project.start_report:
        report_data['proj_days'] = (day.production_date - project.start_report).days + 1

    else:
        report_data['proj_days'] = 1

    # CSR author and comment
    if _week:=day.project.weeklies.filter(week_report_date=day.production_date):
        week = _week.first()
        try:
            author = week.author.name
        except AttributeError:
            author = ''
        report_data['author_table'] = {'CSR': author}
        report_data['comment_table'] = {'Comments': week.csr_week_comment}

    else:
        report_data['author_table'] = {'CSR': ''}
        report_data['comment_table'] = {'Comments': ''}

    proj_start = (
        project.planned_start_date.strftime('%#d %b %Y')
        if project.planned_start_date else NO_DATE_STR
    )

    report_data['project_table'] = {
        'Project': project.project_name,
        'Project VPs': project.planned_vp,
        f'Area (km{SS_2})': project.planned_area,
        'Proj. Start': proj_start,
        'Crew': project.crew_name,
    }

    # proj_stats
    proj_total = totals_prod['proj_total']
    proj_skips = totals_prod['proj_skips']
    if project.planned_vp > 0:
        proj_complete = (proj_total + proj_skips) / project.planned_vp
    else:
        proj_complete = 0

    report_data['proj_stats_table'] = {
        'Recorded VPs': proj_total,
        f'Area (km{SS_2})': project.planned_area * proj_complete,
        'Skip VPs': proj_skips,
        '% Complete': proj_complete,
        'Est. Complete': r_iface.calc_est_completion_date(
            day, AVG_PERIOD, project.planned_vp, proj_complete),
    }

    # hse_stats
    for period in ['week', 'month', 'proj']:
        report_data[f'hse_stats_{period}_table'] = {
            'LTI': totals_hse[f'{period}_lti'],
            'RWC': totals_hse[f'{period}_rwc'],
            'MTC': totals_hse[f'{period}_mtc'],
            'FAC': totals_hse[f'{period}_fac'],
            'NM/ Incidents': totals_hse[f'{period}_incident_nm'],
            'LSR': totals_hse[f'{period}_lsr_violations'],
            'STOP Cards': totals_hse[f'{period}_stop'],
            'Drills': totals_hse[f'{period}_drills'],
            'Audits': totals_hse[f'{period}_audits'],
            'Medevac': totals_hse[f'{period}_medevac'],
        }

    # prod stats
    for period in ['week', 'month', 'proj']:
        report_data[f'prod_stats_{period}_table'] = {
            'Total VPs': totals_prod[f'{period}_total'],
            'Target VPS': totals_prod[f'{period}_ctm'],
            '% Target': totals_prod[f'{period}_appctm'],
            'Rec. Hrs': totals_time[f'{period}_rec_time'],
            'Standby Hrs': totals_time[f'{period}_standby'],
            'Avg VPs/day': totals_prod[f'{period}_avg'],
            'Down Hrs': totals_time[f'{period}_downtime'],
            'Skip VPs': totals_prod[f'{period}_skips'],
            '% Skips': totals_prod[f'{period}_perc_skips'],
        }

    # days prod stats
    report_data['days_prod'] = {
        'header':['Day', 'VPs', 'Rec time', 'VP/ hr', 'Skips',
                  'Layout', 'Pickup', 'Download', 'Charged',
                  'Failure', 'Repair', 'QC field']
    }

    for key, day in reversed(days.items()):
        report_data['days_prod'][key] = [
            day['date'].strftime('%d-%b-%Y'),
            day['prod']['day_total'],
            day['times']['day_rec_time'],
            day['prod']['day_vp_hour'],
            day['prod']['day_skips'],
            day['rcvr']['day_layout'],
            day['rcvr']['day_pickup'],
            day['rcvr']['day_node_download'],
            day['rcvr']['day_node_charged'],
            day['rcvr']['day_node_failure'],
            day['rcvr']['day_node_repair'],
            day['rcvr']['day_qc_field'],
        ]

    report_data['weeks_prod'] = {}
    report_data['weeks_prod']['header'] = report_data['days_prod']['header'].copy()
    report_data['weeks_prod']['header'][0] = 'Week'

    # weeks prod stats
    for key, wk in weeks.items():
        report_data['weeks_prod'][key] = [
            (wk['dates'][0].strftime('%d-%b-%Y') + ' ' +
             wk['dates'][1].strftime('%d-%b-%Y')),
            wk['prod']['week_total'],
            wk['times']['week_rec_time'],
            wk['prod']['week_vp_hour'],
            wk['prod']['week_skips'],
            wk['rcvr']['week_layout'],
            wk['rcvr']['week_pickup'],
            wk['rcvr']['week_node_download'],
            wk['rcvr']['week_node_charged'],
            wk['rcvr']['week_node_failure'],
            wk['rcvr']['week_node_repair'],
            wk['rcvr']['week_qc_field'],
        ]

    # days times stats
    report_data['days_times'] = {
        'header':['Day', 'Recording','Rec move', 'Logistics', 'WOS', 'WOL',
        'Shift change', 'Company suspension', 'Company requested tests',
        'Beyond contractor control', 'Camp move', 'Rec. eqpmt fault',
        'Vib faults', 'Incidents', 'Legal/ dispute', 'Comp. instruction',
        'Contractor noise', 'Other DT','Ops time', 'Standby', 'Downtime',]
    }

    for key, day in reversed(days.items()):
        report_data['days_times'][key] = [
            day['date'].strftime('%d-%b-%Y'),
            day['times']['day_rec_time'],
            day['times']['day_rec_moveup'],
            day['times']['day_logistics'],
            day['times']['day_wait_source'],
            day['times']['day_wait_layout'],
            day['times']['day_wait_shift_change'],
            day['times']['day_company_suspension'],
            day['times']['day_company_tests'],
            day['times']['day_beyond_control'],
            day['times']['day_camp_move'],
            day['times']['day_rec_eqpmt_fault'],
            day['times']['day_vibrator_fault'],
            day['times']['day_incident'],
            day['times']['day_legal_dispute'],
            day['times']['day_comp_instruction'],
            day['times']['day_contractor_noise'],
            day['times']['day_other_downtime'],
            day['times']['day_ops_time'],
            day['times']['day_standby'],
            day['times']['day_downtime'],
        ]

    #weeks times stat
    report_data['weeks_times'] = {}
    report_data['weeks_times']['header'] = report_data['days_times']['header'].copy()
    report_data['weeks_times']['header'][0] = 'Week'

    for key, wk in weeks.items():
        report_data['weeks_times'][key] = [
            (wk['dates'][0].strftime('%d-%b-%Y') + ' ' +
             wk['dates'][1].strftime('%d-%b-%Y')),
            wk['times']['week_rec_time'],
            wk['times']['week_rec_moveup'],
            wk['times']['week_logistics'],
            wk['times']['week_wait_source'],
            wk['times']['week_wait_layout'],
            wk['times']['week_wait_shift_change'],
            wk['times']['week_company_suspension'],
            wk['times']['week_company_tests'],
            wk['times']['week_beyond_control'],
            wk['times']['week_camp_move'],
            wk['times']['week_rec_eqpmt_fault'],
            wk['times']['week_vibrator_fault'],
            wk['times']['week_incident'],
            wk['times']['week_legal_dispute'],
            wk['times']['week_comp_instruction'],
            wk['times']['week_contractor_noise'],
            wk['times']['week_other_downtime'],
            wk['times']['week_ops_time'],
            wk['times']['week_standby'],
            wk['times']['week_downtime'],
        ]

    # terrain distribution week and project
    report_data['week_terrain'] = {
        'flat': totals_prod[f'week_{source_prod_schema[0][:5]}'],
        'rough': totals_prod[f'week_{source_prod_schema[1][:5]}'],
        'facility': totals_prod[f'week_{source_prod_schema[2][:5]}'],
        'dune': totals_prod[f'week_{source_prod_schema[3][:5]}'],
        'sabkha': totals_prod[f'week_{source_prod_schema[4][:5]}'],
        'skip': totals_prod[f'week_{source_prod_schema[5][:5]}'],
        'total': totals_prod[f'week_total'],
    }
    report_data['proj_terrain'] = {
        'flat': totals_prod[f'proj_{source_prod_schema[0][:5]}'],
        'rough': totals_prod[f'proj_{source_prod_schema[1][:5]}'],
        'facility': totals_prod[f'proj_{source_prod_schema[2][:5]}'],
        'dune': totals_prod[f'proj_{source_prod_schema[3][:5]}'],
        'sabkha': totals_prod[f'proj_{source_prod_schema[4][:5]}'],
        'skip': totals_prod[f'proj_{source_prod_schema[5][:5]}'],
        'total': totals_prod[f'proj_total'],
    }

    return report_data
